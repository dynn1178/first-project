from http.client import CONTINUE
from httplib2 import RelativeURIError
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import selenium.common.exceptions
import time
import datetime
from datetime import date
from bs4 import BeautifulSoup
import openpyxl
from dateutil.relativedelta import *
import requests
from selenium.webdriver.support import expected_conditions as EC
import re
import random

user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.141 Whale/3.15.136.29 Safari/537.36"
options = webdriver.ChromeOptions()
# options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
options.add_argument('user-agent=' + user_agent)
options.add_argument("lang=ko_KR") # 한국어!


d_today = datetime.date.today()

# a = 0

# wb = openpyxl.load_workbook("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx")
# sheet = wb['참좋은']
# list = ['APP2091', 'JPP022']
# arrv = ['동남아', '일본']

# for a in range(len(list)): # list 수 만큼 반복
#     driver = webdriver.Chrome("d:\\업무폴더\\3.Python\\chromedriver.exe", options=options)
#     driver.get("https://m.verygoodtour.com/Product/PackageMaster?MasterCode="+list[a]+"&dType=L&Calendar=Y")
#     time.sleep(random.uniform(4,6))

#     for i in range(4):  # 개월 수 
#         html = driver.page_source
#         soup = BeautifulSoup(html, "html.parser")
#         res = soup.find_all(class_="tours_info_wrap sky")
        
#         for b in range(len(res)):  # 긁어오는 항목 
#             date_n = res[b].find("span", attrs={"class":"first_line"}).get_text().replace(".","-")[:-3] # 날짜
#             title_n = res[b].find("div", attrs={"class":"tour_tit"}).get_text().strip() # 제목
#             period_n = res[b].find("span", attrs={"class":"period"}).get_text().strip().replace(" ","") # 패턴
#             price_n = res[b].find("div", attrs={"class":"price"}).get_text().strip().replace(",","")[:-2] # 가격
#             air_n = res[b].find("p", attrs={"class":"airline"}).get_text().strip().replace("\n","").replace("  ","(").replace(" ","")+")" # 항공
#             air_m = air_n[:2] # 항공2
#             status_n = res[b].find("div", attrs={"class":"status"}).get_text().strip() #상태
#             sheet.append([d_today, list[a], arrv[a], date_n, title_n, period_n, int(price_n), air_m, status_n]) # 엑셀로 넘기기 
#         wb.save("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx") # 엑셀 저장

#         driver.find_element(By.XPATH, "//*[@id='content']/div/div[2]/div[1]/div/button[2]").click() #다음 페이지로 넘김
#     time.sleep(random.uniform(3,4))




# a = 0

# tm_first = datetime.date(d_today.year, d_today.month, 1)
# nm_first = tm_first + relativedelta(months=1)
# tm_last = nm_first + relativedelta(days=-1)

# wb = openpyxl.load_workbook("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx")
# sheet = wb['하나투어']
# list2 = ['MAV1041','MJS1035']
# arrv2 = ['동남아','일본']

# for a in range(len(list2)): # list 수 만큼 반복
#     driver = webdriver.Chrome("d:\\업무폴더\\3.Python\\chromedriver.exe", chrome_options=options)
#     time.sleep(random.uniform(3,5))

#     for i in range(4):  # 개월 수 
#         start_date = str((tm_first + relativedelta(months=i))).replace("-","")
#         end_date = str((tm_last + relativedelta(months=i))).replace("-","")
#         driver.get("https://m.hanatour.com/trp/pkg/CHPC0PKG0119M100?depCityCd=JCN&strtDepDay="+start_date+"&endDepDay="+end_date+"&rprsProdCds="+list2[a]+"&inpPathCd=DCM&pageSize=20&page=1&prePage=CHPC0PKG0100M100")
#         time.sleep(random.uniform(4,6))
#         while True:
#             try:
#                 driver.find_element(By.XPATH, "//*[@id='container']/div/div[2]/div/div[2]/div/div/a/span").click()
#                 time.sleep(random.uniform(3,4))
#             except:
#                 break
                
#         html = driver.page_source
#         soup = BeautifulSoup(html, "html.parser")
#         res = soup.find_all(class_="item_unit")
                
#         for b in range(len(res)):  # 긁어오는 항목 
#             date_n = res[b].find("span", attrs={"class":"item cal"}).get_text().strip().replace(".","-")[:5]
#             title_n = res[b].find("p", attrs={"item_title eps2"}).get_text().strip() # 제목
#             period_n = res[b].find("span", attrs={"class":"item cal2"}).get_text().strip().replace(" ","") # 패턴
#             price_n = res[b].find("strong", attrs={"class":"price now"}).get_text().strip().replace(",","").replace("원","") # 가격
#             air_n = res[b].find("span", attrs={"class":"air_name ico pl0"}).get_text().strip() # 항공
#             air_m = air_n
#             status_n = res[b].find("span", attrs={"class":"attr brand13"}).get_text().strip() #상태
#             sheet.append([d_today, list2[a], arrv[a], date_n, title_n, period_n, int(price_n), air_m, status_n]) # 엑셀로 넘기기 
#         wb.save("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx") # 엑셀 저장
        

    


# a = 0


# wb = openpyxl.load_workbook("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx")
# sheet = wb['모두투어']
# list3 = ['540898', '540812']
# arrv3 = ['동남아', '일본']

# for a in range(len(list3)): # list 수 만큼 반복
#     driver = webdriver.Chrome("d:\\업무폴더\\3.Python\\chromedriver.exe", chrome_options=options)
#     driver.get("https://www.modetour.com/pkg/Item.aspx?MLoc=01&startlocation=ICN&Theme=THE88&idx="+list3[a])
#     time.sleep(random.uniform(4,5))  # 5초 딜레이

#     for i in range(5):  # 개월 수 
#         html = driver.page_source
#         soup = BeautifulSoup(html, "html.parser")
#         res = soup.find(class_="tb_goods")
#         res.find(id="noItem").decompose()   # 태그 제거
#         res.find(id="loadingLayer").decompose()  # 태그 제거
#         res_cnt = res.find_all('tr')[1:]    # 태그 제거 (첫번째)
    
#         for b in range(len(res_cnt)):  # 긁어오는 항목 
#             date_n = res_cnt[b].find("span", attrs={"class":"start"}).get_text()[:5].replace("월","-")
#             title_n = res_cnt[b].find("td", attrs={"name"}).get_text()
#             period_n = res_cnt[b].find("td", attrs={"name"}).previous_sibling.get_text().replace(" ","")
#             price_n = res_cnt[b].find("span", attrs={"class":"state_start"}).get_text().strip().replace(",","").replace("원","")
#             air_n = res_cnt[b].find('img')['alt']
#             air_m = air_n
#             status_n = (res_cnt[b].find("a", attrs={"class":"btn_view"})).parent.get_text()[:4]
#             sheet.append([d_today.strftime("%Y-%m-%d"), int(list3[a]), arrv3[a], date_n, title_n, period_n, int(price_n), air_m, status_n]) # 엑셀로 넘기기 
#             wb.save("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx") # 엑셀 저장
#         time.sleep(random.uniform(1,2))
#         if i == 2:
#             driver.find_element(By.XPATH, "//*[@id='container']/div/div[2]/div/div[1]/div[3]/div[2]/span[2]/a").click() #다음 슬라이스 넘김
#             time.sleep(random.uniform(1,2))
#         if i == 4:
#             break 
#         else:
#             driver.find_element(By.XPATH, f"/html/body/div[4]/div/div[2]/div/div[1]/div[3]/div[1]/div/div[1]/ul/li[{i+1}]").click() #다음 페이지로 넘김
            
#         time.sleep(random.uniform(1,2))



# wb = openpyxl.load_workbook("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx")
# sheet = wb['인터파크']
# list4 = ['A1033468', 'A6013398']
# arrv4 = ['동남아', '일본']

# for a in range(len(list4)): # list 수 만큼 반복
#     driver = webdriver.Chrome("d:\\업무폴더\\3.Python\\chromedriver.exe", chrome_options=options)
#     driver.get("https://tour.interpark.com/goods/detail/?BaseGoodsCd="+list4[a])
#     time.sleep(random.uniform(4,5))  # 5초 딜레이

#     page = driver.find_element(By.XPATH, "//*[@id='dBody']/div[3]/div[1]/div[3]/div[2]/button").text
#     page2 = re.sub("[가-힣]|\(|\)","", page)
#     page3 = re.split('/', page2)
    
#     for i in range(11):
#         driver.find_element(By.XPATH, "//*[@id='dBody']/div[3]/div[1]/div[3]/div[2]/button").click()
#         time.sleep(random.uniform(2,3))
        
#     html = driver.page_source
#     soup = BeautifulSoup(html, "html.parser")
#     res = soup.find("tbody", attrs={"class":"j-GoodsListByDepartureDatetbody"})
#     res2 = res.finㅌd_all('tr')

#     period = (driver.find_element(By.XPATH, "//*[@id='dBody']/div[2]/div[2]/table/tbody/tr[2]/td/strong").text).replace(" ","")
#     title_n = soup.select_one('#dBody > div.default-section.goods-info > h2').get_text()
#     time.sleep(random.uniform(1,2))

#     for b in range(len(res2)):
#         date_n = res2[b].find("td", attrs={"class":"date"}).get_text().replace(".","-")[:5]
#         price_n = res2[b].find("span", attrs={"class":"point01"}).get_text().replace(",","").replace("원","")
#         air_n = (res2[b].find("td", attrs={"class":"date"}).next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling).get_text()
#         air_m = re.sub('[A-Z]|[0-9]','', air_n).strip()
#         status_n = (res2[b].select("span")[2]).text
        
#         # print(d_today.strftime("%Y-%m-%d"), list4[a], arrv4[a], date_n, title_n, period, int(price_n), air_m, status_n)
#         sheet.append([d_today.strftime("%Y-%m-%d"), list4[a], arrv4[a], date_n, title_n, period, int(price_n), air_m, status_n])
#     wb.save("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx") # 엑셀 저장
#     time.sleep(random.uniform(1,2))





wb = openpyxl.load_workbook("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx")
sheet = wb['교원투어']
list5 = ['ATP101220701OZ02']
arrv5 = ['동남아']

for a in range(len(list5)): # list 수 만큼 반복
    driver = webdriver.Chrome("d:\\업무폴더\\3.Python\\chromedriver.exe", chrome_options=options)
    driver.get(f"https://kyowontour.com/goods/goodsEventDetail?tourCode={list5[a]}&menuCode=M010301")
    time.sleep(random.uniform(4,5))  # 5초 딜레이

    driver.find_element(By.XPATH, "//*[@id='btn_differentDate']").click()
    time.sleep(random.uniform(1,2)) 
    pick = driver.find_element(By.XPATH, "//*[@id='tourDatepicker']/div/div/div/div/div[2]/div[25]").click()
    time.sleep(random.uniform(1,2)) 
    pick = driver.find_element(By.XPATH, "//*[@id='tourDatepicker']/div/div/div/div/div[2]/div[25]").click()
    time.sleep(random.uniform(1,2)) 
    
    html = driver.page_source       
    soup = BeautifulSoup(html, "html.parser")
    res = soup.find("div", attrs={"class":"mCSB_container"})
    res2 = res.find_all("tr")
    print(res2)
# 


    # for i in range(11):
    #     driver.find_element(By.XPATH, "//*[@id='dBody']/div[3]/div[1]/div[3]/div[2]/button").click()
    #     time.sleep(random.uniform(2,3))
        
    # html = driver.page_source
    # soup = BeautifulSoup(html, "html.parser")
    # res = soup.find("tbody", attrs={"class":"j-GoodsListByDepartureDatetbody"})
    # res2 = res.finㅌd_all('tr')

    # period = (driver.find_element(By.XPATH, "//*[@id='dBody']/div[2]/div[2]/table/tbody/tr[2]/td/strong").text).replace(" ","")
    # title_n = soup.select_one('#dBody > div.default-section.goods-info > h2').get_text()
    # time.sleep(random.uniform(1,2))

    # for b in range(len(res2)):
    #     date_n = res2[b].find("td", attrs={"class":"date"}).get_text().replace(".","-")[:5]
    #     price_n = res2[b].find("span", attrs={"class":"point01"}).get_text().replace(",","").replace("원","")
    #     air_n = (res2[b].find("td", attrs={"class":"date"}).next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling).get_text()
    #     air_m = re.sub('[A-Z]|[0-9]','', air_n).strip()
    #     status_n = (res2[b].select("span")[2]).text
        
    #     # print(d_today.strftime("%Y-%m-%d"), list4[a], arrv4[a], date_n, title_n, period, int(price_n), air_m, status_n)
    #     sheet.append([d_today.strftime("%Y-%m-%d"), list5[a], arrv5[a], date_n, title_n, period, int(price_n), air_m, status_n])
    # wb.save("D:\\업무폴더\\0. Tableau\\1. data\\price_goods.xlsx") # 엑셀 저장
    # time.sleep(random.uniform(1,2))